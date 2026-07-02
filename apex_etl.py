import os
import glob
import datetime
import pandas as pd
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials

# --- CONFIGURATION ---
DATA_DIR = "apex_retail"
TARGET_EXCEL = "Apex_Master_Ledger.xlsx"
GOOGLE_SHEET_ID = "YOUR_DEMO_SHEET_ID_HERE"  # <-- Replace with actual ID
CREDENTIALS_JSON = "credentials.json"

def normalize_branch(name_str):
    """Data Engineering Concept: Normalizing messy, human-entered strings."""
    s = str(name_str).lower()
    if "north" in s: return "North Store"
    if "south" in s: return "South Store"
    if "east" in s: return "East Store"
    return "External"

def extract_and_clean(file_path):
    """Extracts data safely by bypassing human-readable corporate headers."""
    df = pd.read_excel(file_path, header=5, engine='openpyxl')
    df.columns = df.columns.astype(str).str.strip()
    df = df.dropna(how='all')
    return df

def run_pipeline():
    # --- ERROR CHECK 1: Missing Configuration ---
    if GOOGLE_SHEET_ID == "YOUR_DEMO_SHEET_ID_HERE" or not GOOGLE_SHEET_ID:
        print("\n❌ [CONFIGURATION ERROR] Default Google Sheet ID detected.")
        print("💡 FIX: Please open 'apex_etl.py', locate 'GOOGLE_SHEET_ID' at the top, and paste the ID from your browser's URL bar.")
        return

    sending_data = []
    receiving_data = []
    branches = {"north": "North Store", "south": "South Store", "east": "East Store"}

    print("⏳ [STAGE 1] Extracting and transforming regional silos...")

    for folder, standard_name in branches.items():
        # 1. Extract Sending Data
        for file in glob.glob(os.path.join(DATA_DIR, folder, "sending", "*.xlsx")):
            if os.path.basename(file).startswith("~$"): continue

            df = extract_and_clean(file)
            df = df[df['Receipt Type'] == 'Sales']

            df['From'] = standard_name
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df['Sending#'] = df['Receipt #']
            sending_data.append(df[['From', 'Date', 'Sending#', 'Total']])

        # 2. Extract Receiving Data
        for file in glob.glob(os.path.join(DATA_DIR, folder, "receiving", "*.xlsx")):
            if os.path.basename(file).startswith("~$"): continue

            df = extract_and_clean(file)
            df = df[df['Type'] == 'Receiving']

            df['From_Normalized'] = df['Vendor'].apply(normalize_branch)
            df['To'] = standard_name
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            df['Receiving#'] = df['Voucher #']
            df['StockTransfer#'] = df['Invoice #']
            receiving_data.append(df[['From_Normalized', 'To', 'Date', 'StockTransfer#', 'Receiving#', 'Total']])

    master_send = pd.concat(sending_data, ignore_index=True) if sending_data else pd.DataFrame()
    master_recv = pd.concat(receiving_data, ignore_index=True) if receiving_data else pd.DataFrame()

    if master_send.empty or master_recv.empty:
        print("❌ No data found to process. Did you run setup_demo.py first?")
        return

    print("⏳ [STAGE 2] Reconciling ledger matches (Idempotent Join)...")
    matched = pd.merge(
        master_send, master_recv,
        left_on=['Sending#', 'From'],
        right_on=['StockTransfer#', 'From_Normalized'],
        how='inner'
    )

    for col in ["Sending#", "StockTransfer#", "Receiving#"]:
        matched[col] = matched[col].astype(str).str.replace(r'\.0$', '', regex=True)

    print("⏳ [STAGE 3] Local Backup: Non-destructive Excel injection...")

    try:
        wb = load_workbook(TARGET_EXCEL)
    except FileNotFoundError:
        print(f"\n❌ [FILE ERROR] Could not find {TARGET_EXCEL}.")
        print("💡 FIX: Please run 'python setup_demo.py' to generate the mock environment first.")
        return

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tab_matches = matched[matched['To'] == sheet_name].copy()

        existing_vouchers = {str(ws.cell(row=r, column=9).value) for r in range(1, ws.max_row + 1)}
        tab_matches = tab_matches[~tab_matches['Receiving#'].isin(existing_vouchers)]

        if tab_matches.empty: continue

        append_row = max((r for r in range(ws.max_row, 0, -1) if ws.cell(row=r, column=1).value), default=3) + 1
        for _, row in tab_matches.iterrows():
            sent_dt = row['Date_x'].strftime('%Y-%m-%d') if pd.notna(row['Date_x']) else ""
            recv_dt = row['Date_y'].strftime('%Y-%m-%d') if pd.notna(row['Date_y']) else ""

            record = [row['From'], sent_dt, row['Sending#'], "Sending", row['Total_x'],
                      row['To'], recv_dt, row['StockTransfer#'], row['Receiving#'], "Receiving", row['Total_y']]
            for col_idx, val in enumerate(record, start=1):
                ws.cell(row=append_row, column=col_idx, value=val)
            append_row += 1

        print(f"   ✅ Injected {len(tab_matches)} verified transfers into '{sheet_name}'.")

    # --- ERROR CHECK 2: Locked Excel File ---
    try:
        wb.save(TARGET_EXCEL)
    except PermissionError:
        print("\n❌ [SYSTEM ERROR] Permission Denied while saving the Excel file.")
        print("💡 FIX: Never Leave Excel Open in Edit Mode!")
        print("   If you click inside an Excel cell during your demo and leave your cursor blinking there, Windows locks the file.")
        print("   Close Excel entirely or click outside the cell, then run this script again.")
        return

    print("⏳ [STAGE 4] Cloud Sync: Pushing data to Google Sheets APIs...")

    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(CREDENTIALS_JSON, scopes=scopes)
        client = gspread.authorize(creds)

        # We wrap the entire sheet generation and data appending process in the API try block
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)

        for sheet_name in branches.values():
            cloud_matches = matched[matched['To'] == sheet_name].copy()

            try:
                worksheet = spreadsheet.worksheet(sheet_name)
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols=15)
                worksheet.update('A1:K3', [["Apex Corporate Ledger"], ["Sending Info", "", "", "", "", "Receiving Info"], ['From', 'SentOn', 'Sending#', 'OperationType', 'AmountSent', 'To', 'ReceivedOn', 'StockTransfer#', 'Receiving #', 'OperationType', 'AmountReceived']])

            gs_data = worksheet.get_all_values()
            cloud_existing_vouchers = {str(row[8]).split('.')[0].strip() for row in gs_data[3:] if len(row) > 8}
            cloud_matches = cloud_matches[~cloud_matches['Receiving#'].isin(cloud_existing_vouchers)]

            if cloud_matches.empty:
                print(f"   📊 Cloud '{sheet_name}' is up to date.")
                continue

            records_to_append = cloud_matches[['From', 'Date_x', 'Sending#', 'Total_x', 'To', 'Date_y', 'StockTransfer#', 'Receiving#', 'Total_y']].copy()
            records_to_append.insert(3, 'Op1', 'Sending')
            records_to_append.insert(9, 'Op2', 'Receiving')

            records_to_append = records_to_append.map(lambda x: str(x) if isinstance(x, (datetime.datetime, datetime.date, pd.Timestamp)) else x).fillna("")

            payload = records_to_append.values.tolist()
            worksheet.append_rows(payload, value_input_option='USER_ENTERED')
            print(f"   🚀 Synced {len(payload)} verified transfers to Google Sheets -> '{sheet_name}'.")

        print("🏁 Pipeline execution complete! Zero-dollar data stack fully synchronized.")

    except FileNotFoundError:
        print("\n❌ [AUTH ERROR] Could not find credentials.json.")
        print("💡 FIX: Ensure you placed your Service Account JSON file in this directory and renamed it exactly to 'credentials.json'.")
        return
    except gspread.exceptions.SpreadsheetNotFound:
        print("\n❌ [API ERROR] 404: Google Sheet Not Found.")
        print("💡 FIX 1: Ensure you copied the exact ID from your browser's URL bar into GOOGLE_SHEET_ID.")
        print("💡 FIX 2: Did you remember to click 'Share' in Google Sheets and give Editor access to your Service Account email?")
        return
    except gspread.exceptions.APIError as api_err:
        # --- ERROR CHECK 3: The 403 Viewer Permission Trap ---
        if "403" in str(api_err) or "permission" in str(api_err).lower():
            print("\n❌ [PERMISSION ERROR] 403: The caller does not have permission to edit this sheet.")
            print("💡 FIX: You added the Service Account as a 'Viewer'. You must change it to 'Editor'!")
            print("   1. Open your Google Sheet in your browser.")
            print("   2. Click the 'Share' button in the top right.")
            print("   3. Find the 'etl-worker' email and change its role from 'Viewer' to 'Editor'.")
        else:
            print(f"\n❌ [API ERROR] Failed to connect to Google Cloud: {api_err}")
            print("💡 FIX: Make sure you enabled the 'Google Sheets API' inside your Google Cloud Console project.")
        return

if __name__ == "__main__":
    run_pipeline()
